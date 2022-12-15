#1
import openpyxl
# 2
import xlsxwriter
from datetime import datetime
from dateutil.rrule import rrule, DAILY
import calendar
import holidays.countries


#1 read data
# open file xlsx
wookbook = openpyxl.load_workbook("plan.xlsx")
worksheet = wookbook.active

# first cell
col = 3
row = 6

# create inner dict
user_vacation = {}
# create result dict
vacation = {}

# func to create user_vacation dict + save last number of vacation in the dict
def user_vacation_func(data, duration, num=1, shift=0):
    duration = str(duration)
    duration_ = duration.split()
    data = str(data)
    data_ = data.split()
    for n, i in enumerate(data_):
        if (n + 1) % 2 != 1:
            data_.remove(i)
    len_ = int(len(data_) / 2)
    for j in range(0, len_):
        user_vacation[j + 1 + shift] = data_[num - 1], data_[num], duration_[j]
        num += 2
        last_num = j + 1
    return user_vacation, last_num

# read all lines in the file
while row != 26:
    # find data we need
    employee = worksheet.cell(row, col).value
    durations1 = worksheet.cell(row, col + 3).value
    durations2 = worksheet.cell(row, col + 10).value
    data1 = worksheet.cell(row, col + 4).value
    data2 = worksheet.cell(row, col + 11).value
    # if cell is not None
    if durations1 or data1:
        user_vacation_func(data1,durations1)
        last_num = user_vacation_func(data1,durations1)[1]
    else:
        last_num = 0
    # if cell is not None
    if durations2 or data2:
        # add data in the inner dict
        user_vacation_func(data2, durations2, shift=last_num)
    # get a result dict
    vacation[employee.strip()] = user_vacation
    # erase the inner dict
    user_vacation = {}
    # next line
    row += 1
# print(user_vacation)
# print(vacation['Шматович Дмитрий'])

# 2 schedule builder
week_days = { 1: 'Mon', 2: 'Tue', 3: 'Wed', 4: 'Thu', 5: 'Fri', 6: 'Sat', 7: 'Sun' }
year_months = { 1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June', 7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December' }
year = 2023
previous_year = 2022

listOfHolidays = []
listOfHolidays_date = []
for ptr in holidays.RS(years=year).items(): # https://pythonpip.ru/osnovy/modul-holidays-python
    listOfHolidays.append((ptr[0], ptr[1]))
    listOfHolidays_date.append(ptr[0])

# initializing the start and end date
start_date = datetime(year,1,1)
end_date = datetime(year,12,31)

default_row = 4
default_col = 3

# create a file
file_name = xlsxwriter.Workbook(f"{year} schedule.xlsx")
worksheet = file_name.add_worksheet(f"{year} schedule")


# features for the cells

# set a column's width
worksheet.set_row(2, 8)  # 2
# worksheet.set_row(1, 3)  # 2
worksheet.set_column("A:A", 1)  # A
worksheet.set_column("B:B", 5)  # B
worksheet.set_column("C:C", 30) # C

# months
month_format1 = file_name.add_format({'bold': True, 'align': 'center', 'valign': 'vcenter', 'font_size':16, 'bg_color':'#EDEED6', 'border':1, 'border_color':'black'})
month_format2 = file_name.add_format({'bold': True, 'align': 'center', 'font_size':16, 'bg_color':'#F1FFE7', 'border':1, 'border_color':'black'})
# week_days_format
week_days_format = file_name.add_format({'border':1, 'border_color':'black'})
week_days_format_dayoff = file_name.add_format({'border':1, 'border_color':'black', 'bg_color':'#E6B8B7'})
holiday_format = file_name.add_format({'border':1, 'border_color':'black', 'bg_color':'#EE7A6E'})
# days_format
days_format1 = file_name.add_format({'bg_color':'#F2F2F2', 'border':1, 'border_color':'black'})
days_format2 = file_name.add_format({'bg_color':'#E6E6E6', 'border':1, 'border_color':'black'})
# employee_format
employee_format1 = file_name.add_format({'border':1, 'border_color':'black', 'align': 'center'})
employee_format = file_name.add_format({'border':1, 'border_color':'black'})
# vacation_format
vacation_format = file_name.add_format({'border':1, 'bg_color':'#B7DEE8', 'border_color':'black', 'align': 'center'})
# empty_cell_format
empty_cell_format = file_name.add_format({'border':1, 'border_color':'black'})
# signature_format
signature_format = file_name.add_format({'font_size': 9, 'color': 'blue', 'underline': True})

# freeze the space
worksheet.freeze_panes(6,3)

# Company's logo
worksheet.insert_image('B1', 'GST.png')

for d in rrule(DAILY, dtstart=start_date, until=end_date):
    worksheet.write(default_row, default_col, '', empty_cell_format)
    worksheet.write(default_row + 1, default_col, '', empty_cell_format)
    for line_number in range(len(vacation.keys())):  # number of employees
        worksheet.write(default_row + line_number + 2, default_col, '', empty_cell_format)
    default_col += 1

default_row = 4
default_col = 3

# iterating over the dates
for d in rrule(DAILY, dtstart=start_date, until=end_date):
    # print(year_months[d.month], d.day, week_days[(d.weekday()+1)])
    # print(d.day)
    # print(d.date())

    for employee_num, employee__ in enumerate(vacation.keys()):
        # print(employee__, vacation[employee__])
        for vacation_num in vacation[employee__]:
            # print(employee_num+1, employee__, vacation_num, vacation[employee__][vacation_num]) # vacation[employee__][vacation_num] - vacantion

            # set variables to compare
            first_vacation_day = int(vacation[employee__][vacation_num][0][:2])
            first_vacation_month = int(vacation[employee__][vacation_num][0][3:5])
            last_vacation_day = int(vacation[employee__][vacation_num][1][:2])
            last_vacation_month = int(vacation[employee__][vacation_num][1][3:5])
            vacation_duration = int(vacation[employee__][vacation_num][2])

            # find days off number ('Sat' and 'Sun')
            start = datetime(year, first_vacation_month, first_vacation_day)
            end = datetime(year, last_vacation_month, last_vacation_day)
            dayoff_number = 0
            for day in rrule(DAILY, dtstart=start, until=end):
                # print(year_months[d.month], d.day, week_days[(d.weekday() + 1)])
                if week_days[(day.weekday() + 1)] == 'Sat' or week_days[(day.weekday() + 1)] == 'Sun':
                    dayoff_number += 1
                # add official holidays
                if day.date() in listOfHolidays_date and week_days[(day.weekday() + 1)] != 'Sat' and week_days[(day.weekday() + 1)] != 'Sun':
                    dayoff_number += 1

            # create a schedule
            if d.day == first_vacation_day and d.month == first_vacation_month:
               for day_ in range(vacation_duration + dayoff_number):
                    worksheet.write(default_row + 1 + employee_num+1, default_col + day_, vacation_duration, vacation_format)

    # create days
    # set different color for the cells
    if d.month % 2 != 0:
        worksheet.write(default_row + 1, default_col, d.day, days_format1)
    if d.month % 2 == 0:
        worksheet.write(default_row + 1, default_col, d.day, days_format2)

    # create week_days
    if week_days[(d.weekday()+1)] == 'Sat' or week_days[(d.weekday()+1)] == 'Sun':
        worksheet.write(default_row, default_col, week_days[(d.weekday() + 1)], week_days_format_dayoff)
        worksheet.write(default_row+1, default_col, d.day, week_days_format_dayoff)
        for line_number in range(len(vacation.keys())): # number of employees
            worksheet.write(default_row + line_number+2, default_col, '', week_days_format_dayoff)
    # add official holidays
    elif d.date() in listOfHolidays_date:
        worksheet.write(default_row, default_col, week_days[(d.weekday() + 1)], holiday_format)

        worksheet.write(default_row + 1, default_col, d.day, holiday_format)
        for line_number in range(len(vacation.keys())):  # number of employees
            worksheet.write(default_row + line_number+2, default_col, '', holiday_format)
    else:
        worksheet.write(default_row, default_col, week_days[(d.weekday()+1)], week_days_format)

    # add a holiday name
    for dd in listOfHolidays:
        if d.date() == dd[0]:
            worksheet.write_comment(default_row, default_col, dd[1], {'visible': True, 'x_scale': 0.5, 'y_scale': 0.4, 'x_offset': 30, 'y_offset': -35})

    # create months and merge the cells
    if d.day == 1:
        start_point = default_col
        end_point = calendar.monthrange(year, d.month)[1]

        # set different color for the cells
        if d.month % 2 == 0:
            worksheet.merge_range(default_row - 1, start_point, default_row - 1, default_col + end_point-1,
                                  year_months[d.month], month_format2)
        if d.month % 2 != 0:
            worksheet.merge_range(default_row - 1, start_point, default_row - 1, default_col + end_point - 1,
                                      year_months[d.month], month_format1)

    # go to the next day
    default_col += 1

# create employees
worksheet.merge_range(3, 1, 5, 2, "Employee", month_format1)
default_row = 4
default_col = 2
for n_, employee_ in enumerate(vacation.keys()):
    # create an employee
    worksheet.write(default_row + 2 + n_, default_col, employee_, employee_format)
    # employee number
    worksheet.write(default_row + 2 + n_, default_col - 1, n_+1, employee_format1)

# print(default_row + len(vacation.keys()))
worksheet.write(default_row + len(vacation.keys())+3, default_col-1, "Created by @Victor Vetoshkin", signature_format)

# file protecting option
worksheet.protect(options={'autofilter': True})

file_name.close()
